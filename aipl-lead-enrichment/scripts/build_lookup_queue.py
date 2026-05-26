"""
build_lookup_queue.py
=====================
Given the current enriched master file, generate a prioritized 4-tab Excel
showing the team which companies to spend Lusha/Apollo/Signal Hire/Contact
Out credits on. Allocation logic per references/credit-allocation.md.

USAGE (from Claude analysis tool):
    from build_lookup_queue import build_queue

    queue_path = build_queue(
        master_file='/mnt/user-data/uploads/Hygienic_Leads.xlsx',
        output_dir='/mnt/user-data/outputs',
        budgets={'Lusha': 40, 'Apollo': 10, 'Signal Hire': 10, 'Contact Out': 10},
    )
    # Returns: path to Lookup_Queue.xlsx
"""
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ----- Size heuristics from company name -----
LARGE_KEYWORDS = [
    'OIL AND NATURAL GAS', 'ONGC', 'BARODA', 'MOTILAL OSWAL',
    'HINDWARE', 'BATLIBOI', 'APLAB', 'CHEMBOND', 'MANUGRAPH',
    'ANAND RATHI', 'INDIAN EMULSIFIERS', 'TARAPUR TRANSFORMERS',
    'SKY INDUSTRIES', 'EPC CONSTRUCTIONS', 'JOSTS', 'NTPC',
    'BHEL', 'SAIL', 'GAIL', 'IOCL',
]

def _size_tier(company_name):
    """Infer company size from name conventions."""
    n = (company_name or '').upper()
    if any(k in n for k in LARGE_KEYWORDS):
        return 'large'
    # PSU / govt hints
    if 'CORPORATION LIMITED' in n or 'CORPORATION LTD' in n:
        return 'large'
    if 'PRIVATE LIMITED' in n or 'PVT LTD' in n or 'PVT. LTD' in n or 'PVT LIMITED' in n:
        return 'small'
    if 'LIABILITY PARTNERSHIP' in n or 'LLP' in n:
        return 'small'
    if 'SAHAKARI' in n or 'PATSANTHA' in n:
        return 'cooperative'
    if n.endswith('LIMITED') or n.endswith('LTD'):
        return 'mid'
    return 'small'

def _is_it_role(designation):
    if not designation: return False
    return designation in {
        'VP IT / CISO / CTO', 'IT Manager / IT Head',
        'IT Infra / Sr. IT Infra', 'IT Procurement / Purchase',
    } and 'ROLE FLAG' not in designation

def _has_gatekeeper(row):
    """True if row has a name but not an IT-specific role."""
    has_name = bool(str(row.get('First Name','')).strip() or str(row.get('Last Name','')).strip())
    has_flag = 'ROLE FLAG' in str(row.get('Additional Details',''))
    return has_name and has_flag

def _score_company(row):
    """Higher score = higher priority for paid-tool lookup."""
    company = str(row.get('Company',''))
    designation = str(row.get('Designation',''))
    first = str(row.get('First Name','')).strip()
    last  = str(row.get('Last Name','')).strip()
    email = str(row.get('Primary Email','')).strip()
    phone = (str(row.get('Office Phone','')).strip() or
             str(row.get('Mobile Phone','')).strip())
    has_name = bool(first or last)
    tier = _size_tier(company)

    # Already have IT-specific contact + email + phone → skip
    if _is_it_role(designation) and email and phone:
        return (0, 'skip', 'Already fully enriched')

    # Blank rows always get highest priority
    if not has_name:
        return (10, 'blank', 'No contact found — get any decision-maker')

    # Found gatekeeper at large company → high value to find actual IT person
    if _has_gatekeeper(row) and tier == 'large':
        return (9, 'gatekeeper_large', 'Large co — find actual IT Head, not just MD')
    if _has_gatekeeper(row) and tier == 'mid':
        return (6, 'gatekeeper_mid', 'Mid co — find dedicated IT person if exists')
    if _has_gatekeeper(row) and tier == 'small':
        return (2, 'gatekeeper_small', 'Small co — MD is the IT decision-maker (low ROI)')

    # Found name with no email → enrichment fills email
    if has_name and not email:
        return (5, 'missing_email', 'Have name, missing email — get verified email')
    if has_name and not phone:
        return (4, 'missing_phone', 'Have name + email, missing phone — get direct dial')

    # Cooperative society
    if tier == 'cooperative':
        return (1, 'cooperative', 'Not MCA-registered — likely unreachable via tools')

    return (0, 'covered', 'No further enrichment needed')

def _best_tool(category, tier):
    """Match a company category to the best tool."""
    if category == 'blank':
        return 'Lusha' if tier in ('small', 'cooperative') else 'Apollo'
    if category == 'gatekeeper_large':
        return 'Apollo'   # Best for IT-role discovery at scale
    if category == 'gatekeeper_mid':
        return 'Signal Hire'  # Senior decision-makers
    if category == 'missing_email':
        return 'Contact Out'  # Specialized LinkedIn-to-email
    if category == 'missing_phone':
        return 'Lusha'    # Best phone coverage
    return 'Lusha'

def build_queue(master_file, output_dir='/mnt/user-data/outputs',
                budgets=None, filename='Lookup_Queue.xlsx'):
    """
    Build the prioritized lookup queue.

    master_file:  path to the current enriched XLSX (output of build_vtiger_file.py)
    output_dir:   where to write Lookup_Queue.xlsx
    budgets:      dict of {tool_name: credits}; defaults to free-tier limits

    Returns path to the output Excel.
    """
    if budgets is None:
        budgets = {'Lusha': 40, 'Apollo': 10, 'Signal Hire': 10, 'Contact Out': 10}

    df = pd.read_excel(master_file)

    # Score every row
    scored = []
    for _, row in df.iterrows():
        score, category, reason = _score_company(row)
        if score == 0:
            continue  # already covered or skip
        company = str(row.get('Company',''))
        tier = _size_tier(company)
        tool = _best_tool(category, tier)
        # Try to extract LinkedIn from Additional Details
        details = str(row.get('Additional Details',''))
        lk_match = re.search(r'LinkedIn:\s*(\S+)', details)
        linkedin = lk_match.group(1) if lk_match else ''
        scored.append({
            'Rank Score': score,
            'Company': company,
            'City': row.get('City',''),
            'Current Contact': (f"{row.get('First Name','')} {row.get('Last Name','')}".strip() or '(none)'),
            'Current Designation': row.get('Designation',''),
            'What to Look Up': _what_to_lookup(category),
            'Why This Tool': reason,
            'LinkedIn URL': linkedin,
            'Suggested Tool': tool,
            'Size Tier': tier,
        })

    # Sort by score (high first), within score by tier (large first)
    tier_rank = {'large': 0, 'mid': 1, 'small': 2, 'cooperative': 3}
    scored.sort(key=lambda r: (-r['Rank Score'], tier_rank.get(r['Size Tier'], 9)))

    # Allocate to tools based on budgets and suggested tool
    tool_queues = {t: [] for t in budgets}
    overflow = []
    for r in scored:
        t = r['Suggested Tool']
        if t in tool_queues and len(tool_queues[t]) < budgets[t]:
            tool_queues[t].append(r)
        else:
            # Try other tools with remaining capacity
            assigned = False
            for alt in ['Lusha', 'Apollo', 'Signal Hire', 'Contact Out']:
                if alt in tool_queues and len(tool_queues[alt]) < budgets[alt]:
                    r2 = dict(r); r2['Suggested Tool'] = alt
                    r2['Why This Tool'] = f"(originally {t}; reassigned due to credit limits)"
                    tool_queues[alt].append(r2)
                    assigned = True
                    break
            if not assigned:
                overflow.append(r)

    # Build multi-tab Excel
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / filename
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Summary tab first
    ws = wb.create_sheet('Summary')
    ws.append(['Tool', 'Credits Available', 'Companies Queued', '% Used'])
    for cell in ws[1]: cell.font = Font(bold=True)
    for tool, budget in budgets.items():
        n_queued = len(tool_queues[tool])
        ws.append([tool, budget, n_queued, f'{100*n_queued//budget if budget else 0}%'])
    ws.append([])
    ws.append(['Overflow (not assigned — exceed monthly credits):', '', len(overflow), ''])
    ws.append([])
    ws.append(['HOW TO USE THIS FILE:'])
    ws.append(['1. Open each tool tab (Lusha, Apollo, Signal Hire, Contact Out)'])
    ws.append(['2. For each row, look up the company in that tool''s browser extension'])
    ws.append(['3. Unlock the contact, export to CSV'])
    ws.append(['4. When done with all 4 tools, upload all exports back to Claude'])
    ws.append(['5. Say "merge these exports into the master file" — done.'])
    for ci in range(1, 5):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 24

    cols = ['Company','City','Current Contact','Current Designation',
            'What to Look Up','LinkedIn URL','Size Tier','Why This Tool']
    header_fill = PatternFill('solid', start_color='305496')

    for tool, rows in tool_queues.items():
        sheet_name = f'{tool} ({budgets[tool]} credits)'[:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
        for r in rows:
            ws.append([r[c] for c in cols])
        for ci, h in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max(len(h)+2, 14), 50)
        ws.freeze_panes = 'A2'

    if overflow:
        ws = wb.create_sheet('Overflow (exceeds credits)')
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in overflow:
            ws.append([r[c] for c in cols])
        for ci, h in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max(len(h)+2, 14), 50)

    wb.save(out_path)
    return str(out_path)

def _what_to_lookup(category):
    return {
        'blank':              'Any decision-maker — Director/MD/Founder/IT person',
        'gatekeeper_large':   'CIO / CTO / IT Head email + direct phone',
        'gatekeeper_mid':     'IT Manager / IT Head email + phone',
        'gatekeeper_small':   '(low priority) verify MD email/phone',
        'missing_email':      'Direct verified email for the named contact',
        'missing_phone':      'Direct/mobile phone for the named contact',
        'cooperative':        '(unlikely to be in tools) — try office phone',
    }.get(category, 'Any contact info')
