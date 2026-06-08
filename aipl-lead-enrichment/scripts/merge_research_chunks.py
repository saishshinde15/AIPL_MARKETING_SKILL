"""
merge_research_chunks.py
========================
Stitch batched research results back onto a source company list — reliably —
so `build_vtiger_file.build_files()` can find each company's enrichment.

WHY THIS EXISTS
---------------
For big lists (or any "no-cheating, real research" run), the right way to enrich
is in BATCHES: research ~10 companies at a time, capture each batch as a dict
keyed by company name, then MERGE all batches onto the original file's exact
names before building the Vtiger output.

The catch: the research key and the source name are rarely byte-identical —
the source file says "H D F C - Ergo", "Universal Somp", "Reliance Capital Ltd.
(Converted Into CIC W.E.F 31-08-2018)"; the research says "HDFC Ergo",
"Universal Sompo", "Reliance Capital". `build_files()` does an EXACT key match,
so without reconciliation those rows come back blank.

`core()` normalises both sides to a comparable token set and `merge()` matches
them — tolerant of legal-form suffixes, parenthetical "(formerly …)" notes,
spaced-out acronyms, minor typos, and TRUE duplicates (the same company listed
twice). It returns enrichment keyed by the EXACT source names, ready for
`build_files()`.

PORTABILITY NOTE (read this)
----------------------------
This module is the REPRODUCIBLE half of the fast batched workflow. In Claude
Code the research batches can be fanned out to parallel sub-agents (fast); in
the Claude.ai app there are no sub-agents, so Claude runs the batches
SEQUENTIALLY in one conversation — same quality/coverage, just slower. Either
way, the merge step below is identical, deterministic, and free.

USAGE (from the skill)
----------------------
    from merge_research_chunks import merge
    # companies: the list you pass to build_files (dicts with Company/EnterpriseName)
    # chunks: dict {name: enr}, OR list of such dicts, OR list of JSON file paths
    enrichment, report = merge(companies, chunks,
                               extra_stop={'insurance', 'life', 'general'})  # sector words
    paths = build_files(companies, enrichment, output_dir=...)
    # `report` is a list of (idx, source_name, matched_key, score) for transparency

CLI (quick match-report against files)
--------------------------------------
    python merge_research_chunks.py --source list.xlsx --name-col 1 \
        --chunks /tmp/results*.json --extra-stop insurance,life,general

Pure standard library (openpyxl only in the CLI path, which the skill already has).
"""
import json
import re
import glob as _glob
from pathlib import Path

# ---- stopwords that are NEVER distinguishing ------------------------------
# Legal forms + filler. NOTE: bare single letters (w/e/f/s/…) are deliberately
# NOT here — stopping them corrupts spaced acronyms like "H D F C" -> "HDC".
_LEGAL = {
    'private', 'pvt', 'pvt.', 'limited', 'ltd', 'ltd.', 'llp', 'cic',
    'co', 'co.', 'company', 'corporation', 'corp', 'corp.',
    'india', 'indian', 'the', 'and', '&', 'of', 'm/s', 'm', 's', 'ms',
}
# Rename / re-registration noise the source file carries inside the name string.
_RENAME = {
    'formerly', 'previously', 'known', 'as', 'earlier', 'old', 'name',
    'erstwhile', 'converted', 'into', 'to', 'now', 'wef',
}
_DEFAULT_STOP = _LEGAL | _RENAME


def core(name, extra_stop=frozenset()):
    """
    Normalise a company name to a comparable list of distinctive tokens.

    - strips ()/[]/{} notes ("(formerly X)", "(CIC)", "(Tata group)")
    - strips "w.e.f" and dd-mm-yyyy effective-date noise
    - drops legal-form + rename-filler stopwords (+ caller's `extra_stop`,
      e.g. a sector word that's on EVERY row like "Insurance")
    - collapses runs of single letters: "h d f c ergo" -> ["hdfc", "ergo"]
    """
    s = (name or '').lower()
    s = re.sub(r'\(.*?\)|\[.*?\]|\{.*?\}', ' ', s)                  # bracketed notes
    s = re.sub(r'w\.?e\.?f\.?', ' ', s)                            # "w.e.f"
    s = re.sub(r'\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4}', ' ', s)       # dates 31-08-2018
    s = re.sub(r'[^a-z0-9 ]', ' ', s)                              # punctuation
    stop = _DEFAULT_STOP | {w.lower() for w in extra_stop}
    toks = [t for t in s.split() if t not in stop and not (t.isdigit() and len(t) > 4)]
    # collapse consecutive single-letter tokens into one acronym token
    out, buf = [], []
    for t in toks:
        if len(t) == 1:
            buf.append(t)
        else:
            if buf:
                out.append(''.join(buf)); buf = []
            out.append(t)
    if buf:
        out.append(''.join(buf))
    return out


def _tok_sim(a, b):
    """Token-set similarity tolerant of prefixes/typos (somp~sompo, aegeon~aegon)."""
    sa, sb = set(a), set(b)
    if not sa or not sb:
        return 0.0

    def hit(x, S):
        if x in S:
            return True
        for y in S:
            if len(x) >= 4 and len(y) >= 4 and (y.startswith(x) or x.startswith(y)):
                return True
            if (abs(len(x) - len(y)) <= 1 and len(x) >= 5
                    and sum(1 for i in range(min(len(x), len(y))) if x[i] != y[i]) <= 2):
                return True
        return False

    inter = sum(1 for x in sa if hit(x, sb))
    return inter / len(sa | sb)


def _contains(a, b):
    """True if the smaller token set is fully contained in the larger."""
    sa, sb = set(a), set(b)
    small, big = (sa, sb) if len(sa) <= len(sb) else (sb, sa)
    return bool(small) and small <= big


def _company_name(item):
    if isinstance(item, str):
        return item.strip()
    return str(item.get('EnterpriseName') or item.get('Company') or '').strip()


def load_chunks(chunks):
    """
    Accept: a dict {name: enr}; a list of such dicts; or a list/glob of JSON
    file paths (each file a dict {name: enr}). Returns one merged dict.
    Warns (returns count) on identical-string key collisions = true duplicates
    that would otherwise be silently overwritten.
    """
    merged, collisions = {}, 0
    if isinstance(chunks, dict):
        return dict(chunks), 0
    items = []
    for c in chunks:
        if isinstance(c, dict):
            items.append(c)
        else:  # path or glob
            for p in sorted(_glob.glob(str(c))) or [str(c)]:
                if Path(p).exists():
                    items.append(json.load(open(p)))
    for d in items:
        for k, v in d.items():
            if k in merged:
                collisions += 1
            merged[k] = v
    return merged, collisions


def merge(companies, chunks, extra_stop=frozenset(), threshold=0.5):
    """
    Map each source company to its researched enrichment record.

    Returns (enrichment, report):
      enrichment : dict keyed by EXACT source company name -> enrichment dict
                   (ready to hand straight to build_files()).
      report     : list of (idx, source_name, matched_key|'*** UNMATCHED ***', score)

    Matching, in order: exact-core -> prefix/typo similarity -> containment.
    A second pass allows re-using an already-matched record for TRUE duplicate
    source rows (same company listed twice).
    """
    merged, _ = load_chunks(chunks) if not isinstance(chunks, dict) else (dict(chunks), 0)
    if isinstance(chunks, dict):
        merged = dict(chunks)
    index = [(core(k, extra_stop), k, v) for k, v in merged.items()]

    report, used, pending = [], set(), []

    def find(ck, allow_used):
        ckey = ' '.join(ck)
        for toks, k, v in index:                       # 1) exact core
            if (allow_used or k not in used) and ' '.join(toks) == ckey:
                return (1.0, k, v)
        scored = sorted(((_tok_sim(ck, toks), k, v) for toks, k, v in index   # 2) similarity
                         if (allow_used or k not in used)), key=lambda x: -x[0])
        if scored and scored[0][0] >= threshold:
            return scored[0]
        for toks, k, v in index:                       # 3) containment
            if (allow_used or k not in used) and _contains(ck, toks):
                return (0.6, k, v)
        return None

    enrichment = {}
    for idx, item in enumerate(companies, 1):
        name = _company_name(item)
        ck = core(name, extra_stop)
        best = find(ck, allow_used=False)
        if best:
            used.add(best[1]); enrichment[name] = best[2]
            report.append((idx, name, best[1], round(best[0], 2)))
        else:
            report.append((idx, name, '*** UNMATCHED ***', 0.0))
            pending.append((idx, name, ck))

    for idx, name, ck in pending:                       # 2nd pass: true duplicates
        best = find(ck, allow_used=True)
        if best:
            enrichment[name] = best[2]
            for j, (ridx, rn, rk, rs) in enumerate(report):
                if ridx == idx:
                    report[j] = (idx, name, best[1] + ' (dup)', round(best[0], 2)); break

    return enrichment, report


def _selftest():
    """Guard the real edge cases this module was built to survive."""
    companies = [
        'H D F C - Ergo General Insurance Company Ltd',          # spaced acronym
        'Universal Somp General Insurance',                       # typo: Somp(o)
        'Bandhan Life Insurance Limited formaly Aegeon life',     # typo + rename filler
        'Reliance Capital Ltd. (Converted Into CIC W.E.F 31-08-2018)',  # date/paren noise
        'BAJAJ ALLIANZ LIFE INSURANCE COMPANY LIMITED',          # duplicate #1
        'BAJAJ ALLIANZ LIFE INSURANCE COMPANY LIMITED ',         # duplicate #2 (trailing sp)
    ]
    chunks = {
        'HDFC Ergo General Insurance Company Ltd': {'first': 'Sriram', 'designation': 'CTO'},
        'Universal Sompo General Insurance': {'first': 'Vikram', 'designation': 'CTO'},
        'Bandhan Life Insurance (formerly Aegon Life)': {'first': 'Sumanta', 'designation': 'CTO'},
        'Reliance Capital': {'first': 'X', 'designation': 'CISO'},
        'Bajaj Allianz Life Insurance Company Limited': {'first': 'Goutam', 'designation': 'CDIO'},
    }
    enr, report = merge(companies, chunks, extra_stop={'insurance', 'life', 'general'})
    ok = True
    checks = [
        ('H D F C - Ergo General Insurance Company Ltd', 'Sriram'),
        ('Universal Somp General Insurance', 'Vikram'),
        ('Bandhan Life Insurance Limited formaly Aegeon life', 'Sumanta'),
        ('Reliance Capital Ltd. (Converted Into CIC W.E.F 31-08-2018)', 'X'),
        ('BAJAJ ALLIANZ LIFE INSURANCE COMPANY LIMITED', 'Goutam'),
    ]
    for name, want in checks:
        got = enr.get(name, {}).get('first')
        flag = 'OK ' if got == want else 'FAIL'
        if got != want:
            ok = False
        print(f'  [{flag}] {name[:46]:<46} -> {got}')
    matched = sum(1 for _, _, k, _ in report if not k.startswith('***'))
    print(f'  matched {matched}/{len(companies)} (incl. duplicate row)')
    print('SELF-TEST:', 'PASS' if ok and matched == len(companies) else 'FAIL')
    return ok


if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser(description='Merge batched research chunks onto a source company list.')
    ap.add_argument('--source', help='source xlsx/csv with the company list')
    ap.add_argument('--name-col', type=int, default=1, help='0-based column index holding the company name')
    ap.add_argument('--sheet', default=None, help='sheet name (xlsx); default = first sheet')
    ap.add_argument('--header-row', type=int, default=0, help='row index that is the header (data starts after)')
    ap.add_argument('--chunks', nargs='*', default=[], help='JSON chunk files (globs ok)')
    ap.add_argument('--extra-stop', default='', help='comma-separated sector words on every row (e.g. insurance,life)')
    ap.add_argument('--selftest', action='store_true')
    args = ap.parse_args()

    if args.selftest or not args.source:
        _selftest()
        raise SystemExit(0)

    extra = {w.strip().lower() for w in args.extra_stop.split(',') if w.strip()}
    import openpyxl
    wb = openpyxl.load_workbook(args.source, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    companies = [str(r[args.name_col]).strip() for r in rows[args.header_row + 1:]
                 if r[args.name_col] and str(r[args.name_col]).strip()]
    _, report = merge(companies, args.chunks, extra_stop=extra)
    matched = sum(1 for _, _, k, _ in report if not k.startswith('***'))
    print(f'Matched {matched}/{len(companies)}')
    for idx, name, k, sc in report:
        if k.startswith('***') or sc < 1.0:
            print(f'  {idx:4d}. {name[:48]:<48} -> {k} ({sc})')
